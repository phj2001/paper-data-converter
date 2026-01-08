"""
Excel写入器模块
提供统一的Excel文件写入功能，支持命令行和GUI版本
"""

from typing import List, Callable, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class ExcelWriter:
    """统一的Excel写入器"""

    def __init__(self, output_file: str, headers: List[str], progress_callback: Optional[Callable[[float], None]] = None):
        """
        初始化Excel写入器

        Args:
            output_file: 输出文件路径
            headers: 列标题列表（会自动添加"图片名称"列）
            progress_callback: 进度回调函数（可选）
        """
        self.output_file = output_file
        self.headers = headers + ["图片名称"]
        self.progress_callback = progress_callback

        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "数据"

        # 样式定义
        self.header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        self.header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        self.header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.cell_alignment = Alignment(horizontal='left', vertical='center')
        self.thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )

        # 写入表头
        self._write_headers()

        self.current_row = 2  # 数据从第2行开始

    def _write_headers(self):
        """写入表头"""
        for col_idx, header in enumerate(self.headers, start=1):
            cell = self.worksheet.cell(row=1, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.thin_border

    def add_data(self, rows: List[List[str]], image_name: str):
        """添加数据行"""
        for row_data in rows:
            # 写入数据列
            for col_idx, value in enumerate(row_data, start=1):
                cell = self.worksheet.cell(row=self.current_row, column=col_idx, value=value)
                cell.alignment = self.cell_alignment
                cell.border = self.thin_border

            # 写入图片名称（最后一列）
            last_col = len(self.headers)
            cell = self.worksheet.cell(row=self.current_row, column=last_col, value=image_name)
            cell.alignment = self.cell_alignment
            cell.border = self.thin_border

            self.current_row += 1

    def auto_adjust_width(self):
        """自动调整列宽"""
        for col in range(1, len(self.headers) + 1):
            max_length = 0
            column = get_column_letter(col)

            for row in range(1, min(51, self.worksheet.max_row + 1)):
                cell = self.worksheet[f"{column}{row}"]
                if cell.value:
                    cell_value = str(cell.value)
                    chinese_count = sum(1 for c in cell_value if '\u4e00' <= c <= '\u9fff')
                    width = (len(cell_value) + chinese_count) * 1.2
                    max_length = max(max_length, min(width, 50))

            adjusted_width = min(max(10, max_length + 2), 60)
            self.worksheet.column_dimensions[column].width = adjusted_width

    def save(self):
        """保存Excel文件"""
        self.auto_adjust_width()
        self.worksheet.freeze_panes = 'A2'
        self.workbook.save(self.output_file)
